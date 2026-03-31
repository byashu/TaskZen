import re
import io
from datetime import date, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST

from .models import Task, Category, SubTask


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _validate_username(u_name):
    """Return error string or None if valid."""
    if ' ' in u_name:
        return "Username must not contain spaces."
    if u_name and u_name[0].isdigit():
        return "Username must not start with a number."
    return None


def _validate_password(u_pass):
    """Return error string or None if valid."""
    if not re.search(r'[A-Z]', u_pass):
        return "Password must contain at least one uppercase letter."
    if not re.search(r'\d', u_pass):
        return "Password must contain at least one number."
    if not re.search(r'[!@#$%^&*()_+\-=\[\]{};\':"\\|,.<>\/?`~]', u_pass):
        return "Password must contain at least one special character."
    return None


# ---------------------------------------------------------------------------
# 1. CORE PAGES
# ---------------------------------------------------------------------------

def index_view(request):
    return render(request, 'index.html')


@login_required(login_url='login')
def taskboard_view(request):
    tasks = Task.objects.filter(user=request.user)

    search_query = request.GET.get('search')
    if search_query:
        tasks = tasks.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    sort_by = request.GET.get('sort')
    if sort_by == 'priority':
        tasks = tasks.order_by('priority')
    elif sort_by == 'due':
        tasks = tasks.order_by('due_date')
    else:
        tasks = tasks.order_by('-created_at')

    pending_tasks = tasks.filter(completed=False)
    completed_tasks = tasks.filter(completed=True)

    # Notifications: tasks due tomorrow
    tomorrow = date.today() + timedelta(days=1)
    due_tomorrow = Task.objects.filter(
        user=request.user, completed=False, due_date=tomorrow
    )

    categories = Category.objects.filter(user=request.user)

    context = {
        'pending_tasks': pending_tasks,
        'completed_tasks': completed_tasks,
        'search_query': search_query,
        'due_tomorrow': due_tomorrow,
        'due_tomorrow_count': due_tomorrow.count(),
        'categories': categories,
    }
    return render(request, 'taskboard.html', context)


@login_required(login_url='login')
def task_detail(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    subtasks = task.subtasks.all().order_by('created_at')
    return render(request, 'detail.html', {'task': task, 'subtasks': subtasks})


# ---------------------------------------------------------------------------
# 2. SMART FEATURES
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def productivity_stats(request):
    tasks = Task.objects.filter(user=request.user)
    context = {
        'total': tasks.count(),
        'high': tasks.filter(priority='High').count(),
        'medium': tasks.filter(priority='Medium').count(),
        'low': tasks.filter(priority='Low').count(),
        'completed': tasks.filter(completed=True).count(),
        'pending': tasks.filter(completed=False).count(),
    }
    return render(request, 'productivity.html', context)


@login_required(login_url='login')
def task_summary(request):
    tasks = Task.objects.filter(user=request.user).order_by('due_date')
    context = {
        'tasks': tasks,
        'total': tasks.count(),
        'completed': tasks.filter(completed=True).count(),
        'pending': tasks.filter(completed=False).count(),
    }
    return render(request, 'summary.html', context)


# ---------------------------------------------------------------------------
# 3. TASK CRUD
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def add_task(request):
    categories = Category.objects.filter(user=request.user)
    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description')
        priority = request.POST.get('priority')
        due_date_str = request.POST.get('due_date')
        category_id = request.POST.get('category')

        # Validate due date is not in the past
        if due_date_str:
            due_date = date.fromisoformat(due_date_str)
            if due_date < date.today():
                messages.error(request, "Due date cannot be in the past.")
                return render(request, 'add.html', {'categories': categories})
        else:
            due_date = None

        category = None
        if category_id:
            category = Category.objects.filter(pk=category_id, user=request.user).first()

        Task.objects.create(
            user=request.user,
            title=title,
            description=description,
            priority=priority,
            category=category,
            due_date=due_date,
        )
        messages.success(request, "Task added successfully!")
        return redirect('taskboard')
    return render(request, 'add.html', {'categories': categories})


@login_required(login_url='login')
def edit_task(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    categories = Category.objects.filter(user=request.user)
    if request.method == "POST":
        due_date_str = request.POST.get('due_date')
        if due_date_str:
            due_date = date.fromisoformat(due_date_str)
            if due_date < date.today():
                messages.error(request, "Due date cannot be in the past.")
                return render(request, 'edit.html', {'task': task, 'categories': categories})
        else:
            due_date = None

        category_id = request.POST.get('category')
        category = None
        if category_id:
            category = Category.objects.filter(pk=category_id, user=request.user).first()

        task.title = request.POST.get('title')
        task.description = request.POST.get('description')
        task.priority = request.POST.get('priority')
        task.category = category
        task.due_date = due_date
        task.save()
        messages.success(request, "Task updated!")
        return redirect('taskboard')
    return render(request, 'edit.html', {'task': task, 'categories': categories})


@login_required(login_url='login')
@require_POST
def toggle_task(request, pk):
    """Toggle completed status via checkbox on taskboard."""
    task = get_object_or_404(Task, pk=pk, user=request.user)
    task.completed = not task.completed
    task.save()
    return redirect(request.META.get('HTTP_REFERER', 'taskboard'))


@login_required(login_url='login')
def mark_done(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    task.completed = True
    task.save()
    messages.success(request, f"Task '{task.title}' marked as done!")
    return redirect('taskboard')


@login_required(login_url='login')
def delete_task(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    if request.method == "POST":
        task.delete()
        messages.warning(request, "Task deleted successfully.")
    return redirect('taskboard')


# ---------------------------------------------------------------------------
# 4. SUBTASKS
# ---------------------------------------------------------------------------

@login_required(login_url='login')
@require_POST
def add_subtask(request, pk):
    task = get_object_or_404(Task, pk=pk, user=request.user)
    title = request.POST.get('title', '').strip()
    if title:
        SubTask.objects.create(task=task, title=title)
        messages.success(request, "Sub-task added.")
    return redirect('task_detail', pk=pk)


@login_required(login_url='login')
@require_POST
def toggle_subtask(request, pk):
    subtask = get_object_or_404(SubTask, pk=pk, task__user=request.user)
    subtask.completed = not subtask.completed
    subtask.save()
    return redirect('task_detail', pk=subtask.task.pk)


@login_required(login_url='login')
@require_POST
def delete_subtask(request, pk):
    subtask = get_object_or_404(SubTask, pk=pk, task__user=request.user)
    task_pk = subtask.task.pk
    subtask.delete()
    messages.warning(request, "Sub-task deleted.")
    return redirect('task_detail', pk=task_pk)


# ---------------------------------------------------------------------------
# 5. CATEGORIES
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def manage_categories(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            name = request.POST.get('name', '').strip()
            color = request.POST.get('color', '#6f42c1')
            if name:
                Category.objects.create(user=request.user, name=name, color=color)
                messages.success(request, f"Category '{name}' created.")
        elif action == 'delete':
            cat_id = request.POST.get('cat_id')
            Category.objects.filter(pk=cat_id, user=request.user).delete()
            messages.warning(request, "Category deleted.")
        return redirect('manage_categories')

    categories = Category.objects.filter(user=request.user)
    return render(request, 'categories.html', {'categories': categories})


# ---------------------------------------------------------------------------
# 6. EXPORT
# ---------------------------------------------------------------------------

@login_required(login_url='login')
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    tasks = Task.objects.filter(user=request.user).order_by('completed', '-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    headers = ['#', 'Title', 'Description', 'Priority', 'Category', 'Due Date', 'Status', 'Created']
    header_fill = PatternFill(start_color='6f42c1', end_color='6f42c1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=task.title)
        ws.cell(row=row_num, column=3, value=task.description or '')
        ws.cell(row=row_num, column=4, value=task.priority)
        ws.cell(row=row_num, column=5, value=task.category.name if task.category else '')
        ws.cell(row=row_num, column=6, value=str(task.due_date) if task.due_date else '')
        ws.cell(row=row_num, column=7, value='Completed' if task.completed else 'Pending')
        ws.cell(row=row_num, column=8, value=task.created_at.strftime('%Y-%m-%d'))

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="taskzen_tasks.xlsx"'
    return response


@login_required(login_url='login')
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    tasks = Task.objects.filter(user=request.user).order_by('completed', '-created_at')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], textColor=colors.HexColor('#6f42c1'))
    elements = []

    elements.append(Paragraph("TaskZen — My Tasks", title_style))
    elements.append(Paragraph(
        f"Exported on {date.today().strftime('%B %d, %Y')} | User: {request.user.username}",
        styles['Normal']
    ))
    elements.append(Spacer(1, 0.5*cm))

    data = [['#', 'Title', 'Priority', 'Category', 'Due Date', 'Status']]
    for i, task in enumerate(tasks, 1):
        data.append([
            str(i),
            task.title[:40],
            task.priority,
            task.category.name if task.category else '-',
            str(task.due_date) if task.due_date else '-',
            'Done' if task.completed else 'Pending',
        ])

    table = Table(data, colWidths=[1*cm, 6*cm, 2.5*cm, 3*cm, 3*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6f42c1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f0fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="taskzen_tasks.pdf"'
    return response


# ---------------------------------------------------------------------------
# 7. AUTHENTICATION
# ---------------------------------------------------------------------------

def register_view(request):
    if request.method == 'POST':
        u_name = request.POST.get('username', '').strip()
        u_email = request.POST.get('email', '').strip()
        u_pass = request.POST.get('password', '')
        u_conf_pass = request.POST.get('confirm_password', '')

        err = _validate_username(u_name)
        if err:
            messages.error(request, err)
            return redirect('register')

        err = _validate_password(u_pass)
        if err:
            messages.error(request, err)
            return redirect('register')

        if u_pass != u_conf_pass:
            messages.error(request, "Passwords do not match!")
            return redirect('register')

        if User.objects.filter(username=u_name).exists():
            messages.error(request, "Username already taken.")
            return redirect('register')

        User.objects.create_user(username=u_name, email=u_email, password=u_pass)
        messages.success(request, "Account created! Please login.")
        return redirect('login')
    return render(request, 'register.html')


def login_view(request):
    if request.method == 'POST':
        u_name = request.POST.get('username')
        u_pass = request.POST.get('password')
        user = authenticate(request, username=u_name, password=u_pass)
        if user is not None:
            login(request, user)
            return redirect('taskboard')
        else:
            messages.error(request, "Invalid username or password.")
    return render(request, 'login.html')


def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('home')
